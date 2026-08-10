from __future__ import annotations

import datetime as dt
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from typing import Any

from sqlalchemy import select
from sqlalchemy.orm import Session

from src.backend.models import (
    Client,
    DocumentReference,
    Inquiry,
    Project,
    ServiceAgreement,
    SustainabilityMetric,
    TimeEntry,
    Token,
    User,
)


# --------------------------------------------------------------------------- #
# Result container
# --------------------------------------------------------------------------- #
@dataclass
class ImportResult:
    sheet_type: str
    total_rows: int = 0
    created: int = 0
    updated: int = 0
    skipped: int = 0
    errors: list[dict] = field(default_factory=list)

    def add_error(self, row: int, message: str) -> None:
        self.errors.append({"row": row, "message": message})

    @property
    def ok(self) -> bool:
        return not self.errors

    def to_dict(self) -> dict[str, Any]:
        return {
            "sheet_type": self.sheet_type,
            "total_rows": self.total_rows,
            "created": self.created,
            "updated": self.updated,
            "skipped": self.skipped,
            "errors": self.errors,
            "ok": self.ok,
        }


# --------------------------------------------------------------------------- #
# Parsing helpers
# --------------------------------------------------------------------------- #
def _txt(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _parse_date(value: Any) -> dt.date | None:
    if value is None or value == "":
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    s = str(value).strip()
    if s == "" or s.upper() == "N/A":
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"unrecognized date: {s}")


def _parse_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    s = str(value).strip().replace(",", "")
    if s == "" or s.upper() == "N/A":
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        raise ValueError(f"not a number: {s}") from None


def _parse_int(value: Any) -> int | None:
    if value is None:
        return None
    s = str(value).strip().replace(",", "")
    if s == "" or s.upper() == "N/A":
        return None
    try:
        return int(float(s))
    except ValueError:
        raise ValueError(f"not an integer: {s}") from None


def _parse_bool(value: Any) -> bool | None:
    if value is None:
        return None
    s = str(value).strip().lower()
    if s in ("", "n/a"):
        return None
    if s in ("yes", "true", "y", "1"):
        return True
    if s in ("no", "false", "n", "0"):
        return False
    raise ValueError(f"not a boolean: {s}")


# --------------------------------------------------------------------------- #
# FK resolvers
# --------------------------------------------------------------------------- #
def _client_by_code(s: Session, code: str) -> Client | None:
    return s.scalar(select(Client).where(Client.code == code))


def _client_by_name(s: Session, name: str) -> Client | None:
    return s.scalar(select(Client).where(Client.name == name))


def _inquiry_by_ref(s: Session, ref: str) -> Inquiry | None:
    return s.scalar(select(Inquiry).where(Inquiry.reference_id == ref))


def _agreement_by_ref(s: Session, ref: str) -> ServiceAgreement | None:
    return s.scalar(select(ServiceAgreement).where(ServiceAgreement.reference_id == ref))


def _project_by_code(s: Session, code: str) -> Project | None:
    return s.scalar(select(Project).where(Project.code == code))


def _token_by_ref(s: Session, ref: str) -> Token | None:
    return s.scalar(select(Token).where(Token.reference_id == ref))


def _resolve_user_by_name(s: Session, name: str) -> User | None:
    return s.scalar(select(User).where(User.name == name))


def _doc_by_ref(s: Session, reference_id: str) -> DocumentReference | None:
    return s.scalar(
        select(DocumentReference).where(DocumentReference.reference_id == reference_id)
    )


def _ensure_import_user(s: Session) -> User:
    user = s.scalar(select(User).where(User.email == "import@swa.local"))
    if user is None:
        # Attribution-only identity: it must never be able to authenticate.
        # "!" is not a valid bcrypt hash, so verify_password fails closed for any
        # password; is_active=False is the second gate (checked in login and in
        # core.deps); VIEWER keeps the blast radius minimal if it is ever misused.
        user = User(
            email="import@swa.local",
            password_hash="!",
            name="Data Import",
            role=Role.VIEWER,
            is_active=False,
        )
        s.add(user)
        s.flush()
    return user


def _ensure_client(
    s: Session, *, code: str | None, name: str | None
) -> Client | None:
    """Resolve client by code/name; create a minimal stub from sheet data if missing."""
    client = None
    if code:
        client = _client_by_code(s, code)
    if client is None and name:
        client = _client_by_name(s, name)
    if client is not None:
        return client
    if not name and not code:
        return None
    code = code or f"IMP-{(name or 'UNKNOWN')[:20].upper().replace(' ', '-')}"
    name = name or code
    # avoid unique collisions on re-run
    existing = _client_by_code(s, code)
    if existing:
        return existing
    # Synthetic email only when Excel has none — valid format, not a public domain spoof
    slug = "".join(ch if ch.isalnum() else "-" for ch in code.lower()).strip("-")[:40]
    client = Client(
        code=code,
        name=name,
        primary_email=f"import+{slug}@swa.internal",
        country="India",
        client_status="Active",
        is_active=True,
        notes="Created by Excel import because another sheet referenced this client.",
    )
    s.add(client)
    s.flush()
    return client


def _ensure_project(s: Session, code: str, name: str | None = None) -> Project | None:
    if not code:
        return None
    project = _project_by_code(s, code)
    if project is not None:
        return project
    # Hold client for orphan project IDs (e.g. sustainability refs when Project Tracking is empty).
    # Not fake demo data — system staging until full sheets are imported.
    hold = _ensure_client(
        s,
        code="SWA-SYS-UNLINKED",
        name="SWA — unlinked import rows",
    )
    if hold is None:
        return None
    if hold.notes and "unlinked" not in (hold.notes or "").lower():
        hold.notes = (
            "System hold for project/doc IDs referenced in Excel before Project Tracking "
            "rows exist. Re-assign to real clients after full import."
        )
    project = Project(
        code=code,
        client_id=hold.id,
        name=name or code,
        status="Lead",
        description=(
            "Created by Excel import from a cross-sheet reference "
            f"(code {code}). Link to the correct client when Project Tracking is complete."
        ),
        is_active=True,
    )
    s.add(project)
    s.flush()
    return project


# --------------------------------------------------------------------------- #
# Sheet reader
# --------------------------------------------------------------------------- #
def _find_header_index(rows: list[list[Any]], signatures: list[str]) -> int:
    """Pick the *last* matching header in the first 25 rows.

    SWA workbooks often have a template/legend header first, then the real
    column header above the data block (Clients, Tokens, DRN, Time Logging).
    """
    lowered = [sig.lower() for sig in signatures]
    matches: list[int] = []
    for i, row in enumerate(rows[:25]):
        cells = ["" if c is None else str(c).strip().lower() for c in row]
        if all(any(sig in cell for cell in cells) for sig in lowered):
            # Prefer rows that look like short column labels, not "Why it exists" prose
            non_empty = [c for c in cells if c]
            avg_len = sum(len(c) for c in non_empty) / max(len(non_empty), 1)
            if avg_len < 40:
                matches.append(i)
            elif not matches:
                matches.append(i)
    if not matches:
        raise ValueError(f"header row not found (expected columns: {signatures})")
    return matches[-1]


def _record_get(record: dict, *keys: str) -> Any:
    for k in keys:
        if k in record and record[k] not in (None, ""):
            return record[k]
    # case-insensitive fallback
    lower_map = {str(k).strip().lower(): v for k, v in record.items()}
    for k in keys:
        v = lower_map.get(k.lower())
        if v not in (None, ""):
            return v
    return None


def _looks_like_swa_id(value: Any) -> bool:
    s = _txt(value)
    if not s:
        return False
    return s.upper().startswith("SWA-") or bool(
        __import__("re").match(r"^[A-Z]{2,10}-\d", s.upper())
    )


def read_rows(
    file_path: str,
    signatures: list[str],
    key_field: str,
    *,
    alt_key_fields: list[str] | None = None,
    require_swa_key: bool = False,
) -> list[dict]:
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True)
    # Prefer named operational tabs when present
    preferred = ("DRN Sheet", "Project Tracking", "Sheet1")
    ws = None
    for name in preferred:
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]

    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    hdr_idx = _find_header_index(rows, signatures)
    header = [("" if c is None else str(c).strip()) for c in rows[hdr_idx]]
    # Map non-empty header labels to their absolute Excel column index so a blank
    # column A (common in SWA templates) does not shift every field.
    colmap: list[tuple[int, str]] = [(i, h) for i, h in enumerate(header) if h]

    key_fields = [key_field] + list(alt_key_fields or [])

    def _build(cells: list[Any], shift: int) -> dict[str, Any]:
        rec: dict[str, Any] = {}
        for i, name in colmap:
            j = i + shift
            rec[name] = cells[j] if 0 <= j < len(cells) else ""
        return rec

    def _key_of(rec: dict[str, Any]) -> Any:
        for kf in key_fields:
            v = _record_get(rec, kf)
            if v not in (None, ""):
                return v
        return None

    data: list[dict] = []
    for r in rows[hdr_idx + 1 :]:
        if r is None:
            continue
        cells = list(r) if r is not None else []
        record = _build(cells, 0)
        # skip stray duplicate header rows
        cell_texts = [str(v).strip().lower() for v in record.values() if v is not None]
        sig_hits = sum(1 for t in cell_texts if any(sig.lower() == t for sig in signatures))
        if sig_hits >= max(1, len(signatures)):
            continue
        key_val = _key_of(record)

        def _row_score(rec: dict[str, Any]) -> int:
            """Prefer alignment where IDs land on ID columns, not under Type/Author."""
            k = _key_of(rec)
            if not _looks_like_swa_id(k):
                return -1
            score = 1
            assoc = _txt(
                _record_get(rec, "Associated Project/Token ID", "Associated Project ID")
            )
            if assoc and _looks_like_swa_id(assoc):
                score += 3
            doc_type = _txt(_record_get(rec, "Document Type", "Token Type", "Service Name"))
            if doc_type and not _looks_like_swa_id(doc_type):
                score += 1
            author = _txt(_record_get(rec, "Author", "Client Name"))
            if author and not _looks_like_swa_id(author):
                score += 1
            # Doc ref codes often contain CON/DBR/CAS/GAD/DRN/TKN/INQ/CLT/SA
            ku = str(k).upper()
            if any(
                x in ku
                for x in (
                    "-CON-",
                    "-DBR-",
                    "-CAS-",
                    "-GAD-",
                    "-DRN-",
                    "-TKN-",
                    "-INQ-",
                    "-CLT-",
                    "-SA-",
                    "-PRJ-",
                )
            ):
                score += 2
            return score

        # Try shift -1 (data one column left of header labels — DRN sheet)
        shifted = _build(cells, -1)
        if _row_score(shifted) > _row_score(record):
            record, key_val = shifted, _key_of(shifted)
        else:
            key_val = _key_of(record)

        if not _txt(key_val):
            continue
        if require_swa_key and not _looks_like_swa_id(key_val):
            continue
        if isinstance(key_val, str) and len(key_val) > 60:
            continue
        record[key_field] = key_val
        data.append(record)
    return data


# --------------------------------------------------------------------------- #
# Per-sheet importers
# --------------------------------------------------------------------------- #
def _import_clients(s: Session, rows: list[dict], result: ImportResult) -> None:
    for i, d in enumerate(rows, start=2):
        try:
            code = _txt(d.get("Client ID"))
            name = _txt(d.get("Client Name"))
            if not code:
                result.add_error(i, "missing Client ID")
                continue
            if not name:
                result.add_error(i, "missing Client Name")
                continue
            slug = "".join(ch if ch.isalnum() else "-" for ch in code.lower()).strip("-")[:40]
            email = _txt(d.get("Email")) or f"import+{slug}@swa.internal"
            client = _client_by_code(s, code)
            if client is not None:
                client.name = name
                client.primary_email = email
                client.industry = _txt(d.get("Industry"))
                client.primary_phone = _txt(d.get("Phone"))
                client.address = _txt(d.get("Billing Address"))
                client.client_status = _txt(d.get("Client Status")) or "Active"
                client.notes = _txt(d.get("Notes"))
                fi = _txt(d.get("First Inquiry ID"))
                if fi:
                    inq = _inquiry_by_ref(s, fi)
                    if inq is not None:
                        client.first_inquiry_id = inq.id
                    # else: keep null — inquiry may be imported later; not fatal
                result.updated += 1
            else:
                client = Client(
                    code=code,
                    name=name,
                    primary_email=email,
                    industry=_txt(d.get("Industry")),
                    primary_phone=_txt(d.get("Phone")),
                    address=_txt(d.get("Billing Address")),
                    client_status=_txt(d.get("Client Status")) or "Active",
                    notes=_txt(d.get("Notes")),
                    country="India",
                    is_active=True,
                )
                fi = _txt(d.get("First Inquiry ID"))
                if fi:
                    inq = _inquiry_by_ref(s, fi)
                    if inq is not None:
                        client.first_inquiry_id = inq.id
                s.add(client)
                result.created += 1
        except Exception as e:
            result.add_error(i, str(e))


def _import_inquiries(s: Session, rows: list[dict], result: ImportResult) -> None:
    for i, d in enumerate(rows, start=2):
        try:
            reference_id = _txt(d.get("Inquiry ID"))
            if not reference_id:
                result.add_error(i, "missing Inquiry ID")
                continue
            inquiry_date = _parse_date(d.get("Inquiry Date"))
            if inquiry_date is None:
                result.add_error(i, "missing or invalid Inquiry Date")
                continue
            client_name = _txt(d.get("Client Name"))
            status = _txt(d.get("Status")) or "New"
            client = _client_by_name(s, client_name) if client_name else None
            inquiry = _inquiry_by_ref(s, reference_id)
            values = dict(
                inquiry_date=inquiry_date,
                inquiry_type=_txt(d.get("Inquiry Type")),
                inquiry_source=_txt(d.get("Inquiry Source")),
                client_name=client_name or "",
                requirement_summary=_txt(d.get("Requirement Summary")),
                estimated_value=_parse_decimal(d.get("Estimated Value")),
                priority=_txt(d.get("Priority")),
                status=status,
                notes=_txt(d.get("Notes")),
                converted_client_id=client.id if client else None,
            )
            if inquiry is not None:
                for k, v in values.items():
                    setattr(inquiry, k, v)
                result.updated += 1
            else:
                s.add(Inquiry(reference_id=reference_id, **values))
                result.created += 1
        except Exception as e:
            result.add_error(i, str(e))


def _import_agreements(s: Session, rows: list[dict], result: ImportResult) -> None:
    for i, d in enumerate(rows, start=2):
        try:
            reference_id = _txt(d.get("Agreement ID"))
            if not reference_id:
                result.add_error(i, "missing Agreement ID")
                continue
            client = _ensure_client(
                s, code=_txt(d.get("Client ID")), name=_txt(d.get("Client Name"))
            )
            if client is None:
                result.add_error(i, "client not found (Client Name / Client ID)")
                continue
            service_name = _txt(d.get("Service Name"))
            if not service_name:
                result.add_error(i, "missing Service Name")
                continue
            try:
                start_date = _parse_date(d.get("Start Date"))
            except ValueError as e:
                result.add_error(i, str(e))
                continue
            if start_date is None:
                # Real SA samples sometimes omit start date — still import the row
                start_date = dt.date.today()
            inquiry = _inquiry_by_ref(s, _txt(d.get("Inquiry ID"))) if _txt(
                d.get("Inquiry ID")
            ) else None
            agreement = _agreement_by_ref(s, reference_id)
            values = dict(
                client_id=client.id,
                inquiry_id=inquiry.id if inquiry else None,
                service_name=service_name,
                start_date=start_date,
                end_date=_parse_date(d.get("End Date")),
                total_tokens=_parse_int(d.get("Total Tokens")),
                status=_txt(d.get("Status")) or "Active",
                notes=_txt(d.get("Notes")),
            )
            if agreement is not None:
                for k, v in values.items():
                    setattr(agreement, k, v)
                result.updated += 1
            else:
                s.add(ServiceAgreement(reference_id=reference_id, **values))
                result.created += 1
        except Exception as e:
            result.add_error(i, str(e))


def _import_tokens(s: Session, rows: list[dict], result: ImportResult) -> None:
    for i, d in enumerate(rows, start=2):
        try:
            reference_id = _txt(d.get("Token ID"))
            if not reference_id:
                result.add_error(i, "missing Token ID")
                continue
            agreement = _agreement_by_ref(s, _txt(d.get("Agreement ID")))
            if agreement is None:
                result.add_error(i, f"Agreement {_txt(d.get('Agreement ID'))} not found")
                continue
            token_date = _parse_date(d.get("Date"))
            if token_date is None:
                result.add_error(i, "missing or invalid Date")
                continue
            values: dict[str, Any] = dict(
                agreement_id=agreement.id,
                token_date=token_date,
                token_type=_txt(d.get("Token Type")),
                description=_txt(d.get("Description")),
                token_status=_txt(d.get("Token Status")) or "In Progress",
                client_employee_name=_txt(d.get("Client Employee Name")),
            )
            tu = _parse_int(d.get("Tokens Used"))
            if tu is not None:
                values["tokens_used"] = tu
            emp = _txt(d.get("Swa Employee Name/Team Leader")) or _txt(
                d.get("Swa Employee Name")
            )
            if emp:
                u = _resolve_user_by_name(s, emp)
                if u is not None:
                    values["swa_employee_id"] = u.id
            po = _txt(d.get("Project Owner"))
            if po:
                u = _resolve_user_by_name(s, po)
                if u is not None:
                    values["project_owner_id"] = u.id
            token = _token_by_ref(s, reference_id)
            if token is not None:
                for k, v in values.items():
                    setattr(token, k, v)
                result.updated += 1
            else:
                s.add(Token(reference_id=reference_id, **values))
                result.created += 1
        except Exception as e:
            result.add_error(i, str(e))


def _import_document_references(s: Session, rows: list[dict], result: ImportResult) -> None:
    for i, d in enumerate(rows, start=2):
        try:
            reference_id = _txt(_record_get(d, "Doc Ref No", "DRN"))
            if not reference_id:
                result.add_error(i, "missing Doc Ref No / DRN")
                continue
            if not _looks_like_swa_id(reference_id):
                result.skipped += 1
                continue
            assoc = _txt(
                _record_get(d, "Associated Project/Token ID", "Associated Project ID")
            )
            project = _project_by_code(s, assoc) if assoc else None
            token = None
            if project is None and assoc:
                token = _token_by_ref(s, assoc)
                if token is not None and token.project_id is not None:
                    project = s.get(Project, token.project_id)
            if project is None and assoc and _looks_like_swa_id(assoc):
                # Real sample sheets reference project codes not present on Project Tracking
                project = _ensure_project(s, assoc)
            if project is None:
                result.add_error(i, f"Project {assoc} not found")
                continue
            doc_date = _parse_date(d.get("Date"))
            if doc_date is None:
                doc_date = dt.date.today()
            document_type = _txt(d.get("Document Type"))
            if not document_type:
                result.add_error(i, "missing Document Type")
                continue
            values: dict[str, Any] = dict(
                project_id=project.id,
                token_id=token.id if token else None,
                doc_date=doc_date,
                document_type=document_type,
                type_=_txt(d.get("Type")),
                user_ref=_txt(d.get("User")),
                description=_txt(d.get("Description")),
                revision=_txt(d.get("Revision")) or "R0",
                status=_txt(d.get("Status")) or "Draft",
                remarks=_txt(d.get("Remarks")),
            )
            author = _txt(d.get("Author"))
            if author:
                u = _resolve_user_by_name(s, author)
                if u is not None:
                    values["author_id"] = u.id
            docref = _doc_by_ref(s, reference_id)
            if docref is not None:
                for k, v in values.items():
                    setattr(docref, k, v)
                result.updated += 1
            else:
                s.add(DocumentReference(reference_id=reference_id, **values))
                result.created += 1
        except Exception as e:
            result.add_error(i, str(e))


def _import_projects(s: Session, rows: list[dict], result: ImportResult) -> None:
    for i, d in enumerate(rows, start=2):
        try:
            code = _txt(d.get("Project ID"))
            if not code:
                result.add_error(i, "missing Project ID")
                continue
            client = _client_by_code(s, _txt(d.get("Client ID")))
            if client is None:
                result.add_error(i, f"Client {_txt(d.get('Client ID'))} not found")
                continue
            name = _txt(d.get("Project Name"))
            if not name:
                result.add_error(i, "missing Project Name")
                continue
            project = _project_by_code(s, code)
            values = dict(
                client_id=client.id,
                name=name,
                description=_txt(d.get("Notes")),
                status=_txt(d.get("Status Updates")) or "Lead",
                start_date=_parse_date(d.get("Start date")),
                target_end_date=_parse_date(d.get("End date")),
            )
            if project is not None:
                for k, v in values.items():
                    setattr(project, k, v)
                result.updated += 1
            else:
                s.add(Project(code=code, **values))
                result.created += 1
        except Exception as e:
            result.add_error(i, str(e))


def _import_time_logs(s: Session, rows: list[dict], result: ImportResult) -> None:
    user = _ensure_import_user(s)
    for i, d in enumerate(rows, start=2):
        try:
            ref = _txt(d.get("Reference ID"))
            if not ref:
                result.add_error(i, "missing Reference ID")
                continue
            project = _project_by_code(s, ref) if _looks_like_swa_id(ref) else None
            if project is None:
                token = _token_by_ref(s, ref) if _looks_like_swa_id(ref) else None
                docref = (
                    _doc_by_ref(s, ref)
                    if token is None and _looks_like_swa_id(ref)
                    else None
                )
                if token is not None:
                    if token.project_id is not None:
                        project = s.get(Project, token.project_id)
                    else:
                        # bind to stub so hours are not lost
                        project = _ensure_project(
                            s, f"FROM-TOKEN-{token.reference_id}", name=f"Work for {token.reference_id}"
                        )
                        token.project_id = project.id if project else None
                if project is None and docref is not None and docref.project_id is not None:
                    project = s.get(Project, docref.project_id)
                # match by project name column when ref is non-id text
                pname = _txt(d.get("Project Name"))
                if project is None and pname:
                    project = s.scalar(select(Project).where(Project.name == pname))
                if project is None and _looks_like_swa_id(ref):
                    project = _ensure_project(s, ref, name=pname)
                if project is None:
                    result.add_error(i, f"Reference {ref} not found (Project/Token/Doc)")
                    continue
            entry_date = _parse_date(d.get("Date"))
            if entry_date is None:
                result.add_error(i, "missing or invalid Date")
                continue
            hours = _parse_decimal(d.get("Hours Logged"))
            if hours is None:
                result.add_error(i, "missing Hours Logged")
                continue
            description = _txt(d.get("Activity Type")) or "Imported time log"
            remarks = _txt(d.get("Remarks (optional)")) or _txt(d.get("Remarks"))
            if remarks:
                description = f"{description} — {remarks}"
            employee = _txt(d.get("Employee Name"))
            if employee:
                description = f"[{employee}] {description}"
            billable = _parse_decimal(d.get("Billable Hours")) or Decimal("0")
            exists = s.scalar(
                select(TimeEntry).where(
                    TimeEntry.project_id == project.id,
                    TimeEntry.user_id == user.id,
                    TimeEntry.date == entry_date,
                    TimeEntry.description == description,
                    TimeEntry.hours == hours,
                )
            )
            if exists is not None:
                result.skipped += 1
                continue
            s.add(
                TimeEntry(
                    project_id=project.id,
                    user_id=user.id,
                    date=entry_date,
                    hours=hours,
                    description=description,
                    is_billable=billable > 0,
                )
            )
            result.created += 1
        except Exception as e:
            result.add_error(i, str(e))


def _import_sustainability(s: Session, rows: list[dict], result: ImportResult) -> None:
    for i, d in enumerate(rows, start=2):
        try:
            reference_id = _txt(d.get("Reference ID"))
            if not reference_id:
                result.add_error(i, "missing Reference ID")
                continue
            project = _project_by_code(s, reference_id)
            if project is None and _looks_like_swa_id(reference_id):
                project = _ensure_project(s, reference_id)
            if project is None:
                result.add_error(i, f"Project {reference_id} not found")
                continue
            metric = s.scalar(
                select(SustainabilityMetric).where(
                    SustainabilityMetric.reference_id == reference_id
                )
            )
            # Real sheet uses spaced header "Actual / Expected"
            efficiency = _parse_decimal(
                _record_get(
                    d,
                    "Insulation Efficiency (Actual/Expected)",
                    "Insulation Efficiency (Actual / Expected)",
                )
            )
            green_raw = d.get("Compliant with Green Standards")
            green = None
            try:
                green = _parse_bool(green_raw)
            except ValueError:
                # sheet samples use standard names (GRIHA, IGBC) — treat as compliant
                green = bool(_txt(green_raw) and str(green_raw).strip().lower() not in ("no", "n", "false", "0"))
            values = dict(
                project_id=project.id,
                recorded_date=_parse_date(d.get("Date")) or dt.date.today(),
                compliant_with_green_standards=green,
                energy_saved_kwh=_parse_decimal(d.get("Total Energy Saved (kWh)")),
                co2_avoided_tco2e=_parse_decimal(d.get("CO2 emissions avoided (tCO2e)")),
                lifecycle_cost_savings_inr=_parse_decimal(
                    d.get("Lifecycle Cost savings delivered (INR)")
                ),
                insulation_efficiency_ratio=efficiency,
                payback_period_months=_parse_decimal(d.get("Payback Period (Months)")),
                notes=_txt(d.get("Notes")),
            )
            if metric is not None:
                for k, v in values.items():
                    setattr(metric, k, v)
                result.updated += 1
            else:
                s.add(SustainabilityMetric(reference_id=reference_id, **values))
                result.created += 1
        except Exception as e:
            result.add_error(i, str(e))


# --------------------------------------------------------------------------- #
# Dispatch
# --------------------------------------------------------------------------- #
SHEET_CONFIG: dict[str, dict[str, Any]] = {
    "clients": {
        "signatures": ["Client ID", "Client Name"],
        "key_field": "Client ID",
        "require_swa_key": True,
        "fn": _import_clients,
    },
    "inquiries": {
        "signatures": ["Inquiry ID"],
        "key_field": "Inquiry ID",
        "require_swa_key": True,
        "fn": _import_inquiries,
    },
    "agreements": {
        "signatures": ["Agreement ID"],
        "key_field": "Agreement ID",
        "require_swa_key": True,
        "fn": _import_agreements,
    },
    "tokens": {
        "signatures": ["Token ID"],
        "key_field": "Token ID",
        "require_swa_key": True,
        "fn": _import_tokens,
    },
    "document_references": {
        "signatures": ["Doc Ref No"],
        "alt_signatures": ["DRN", "Document Type"],
        "key_field": "Doc Ref No",
        "alt_key_fields": ["DRN"],
        "require_swa_key": True,
        "fn": _import_document_references,
    },
    "projects": {
        "signatures": ["Project ID"],
        "key_field": "Project ID",
        "require_swa_key": True,
        "fn": _import_projects,
    },
    "time_logs": {
        "signatures": ["Hours Logged", "Employee Name"],
        "key_field": "Reference ID",
        "fn": _import_time_logs,
    },
    "sustainability": {
        "signatures": ["Reference ID", "Compliant with Green Standards"],
        "key_field": "Reference ID",
        "require_swa_key": True,
        "fn": _import_sustainability,
    },
}


def import_sheet(
    session: Session, sheet_type: str, file_path: str, commit: bool = False
) -> ImportResult:
    if sheet_type not in SHEET_CONFIG:
        raise ValueError(f"unknown sheet_type: {sheet_type}")
    cfg = SHEET_CONFIG[sheet_type]
    result = ImportResult(sheet_type=sheet_type)
    try:
        try:
            rows = read_rows(
                file_path,
                cfg["signatures"],
                cfg["key_field"],
                alt_key_fields=cfg.get("alt_key_fields"),
                require_swa_key=bool(cfg.get("require_swa_key")),
            )
        except ValueError:
            alt = cfg.get("alt_signatures")
            if not alt:
                raise
            rows = read_rows(
                file_path,
                alt,
                cfg["key_field"],
                alt_key_fields=cfg.get("alt_key_fields"),
                require_swa_key=bool(cfg.get("require_swa_key")),
            )
        result.total_rows = len(rows)
        cfg["fn"](session, rows, result)
        if commit:
            session.commit()
        else:
            session.rollback()
    except Exception as e:
        session.rollback()
        result.add_error(0, f"Fatal: {e}")
    return result
